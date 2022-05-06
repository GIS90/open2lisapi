# -*- coding: utf-8 -*-


"""
------------------------------------------------

describe:
    excel lib

base_info:
    __author__ = "PyGo"
    __time__ = "2022/4/27 04:06 下午"
    __version__ = "v.1.0.0"
    __mail__ = "gaoming971366@163.com"
    __blog__ = "www.pygo2.top"
    __project__ = "open2lisapi"

usage:

design:

reference urls:

python version:
    python3


Enjoy the good life everyday！！!
Life is short, I use python.

------------------------------------------------
"""
import os
import xlwt
import xlrd
import openpyxl
import zipfile

from deploy.utils.utils import filename2md5, \
    get_base_dir, get_now, mk_dirs
from deploy.utils.utils import get_now
from deploy.config import STORE_CACHE
from deploy.utils.status_msg import StatusMsgs
from deploy.utils.enums import *


class ExcelLib(object):

    DEFAULT_PREFIX = '.xlsx'
    DEFAULT_ZIP_PREFIX = '.zip'
    DEFAULT_AUTO_ID = '9999'

    def __init__(self, blank=0):
        """
        class initialize parameters
        :param blank:
        """
        self.prefix_list = ['.xlsx', '.xls']
        self.prefix_zip_list = ['.zip']
        self.blank = 0

    @staticmethod
    def format_res(status_id: int, message: str, data: dict) -> dict:
        """
        方法请求结果格式化
        """
        return {
            'status_id': status_id,
            'message': message if message else StatusMsgs.get(status_id),
            'data': data
        }

    @classmethod
    def font_content(self):
        style_title = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = 'Times New Roman'
        font.bold = True
        font.color_index = 4
        font.height = 220
        style_title.font = font
        return style_title

    @classmethod
    def font_title(self):
        style_content = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = 'Times New Roman'
        font.bold = False
        font.color_index = 4
        font.height = 220
        style_content.font = font
        return style_content

    def read_headers(self, excel_file):
        """
        get the header information of excel file
        :param excel_file: excel file real path
        :return: dict
        result content:
            sheets: dict sheet base info, contain something. such as row.col.index.name
            number sheet: int sheet number
            names: sheets name dict
            columns: sheet columns dict
        """
        res = {'sheets': {}, 'nsheet': 0, 'names': {}, 'columns': {}}
        if not excel_file or not os.path.exists(excel_file):
            return res
        excel = xlrd.open_workbook(excel_file)
        nsheets = excel.nsheets
        # sheet_names = excel.sheet_names()
        sheets_dict = dict()
        names_dict = dict()
        columns_dict = dict()
        for i in range(0, nsheets, 1):
            sheet = excel.sheet_by_index(i)
            names_dict[i] = sheet.name
            sheets_dict[i] = {'row': sheet.nrows, 'col': sheet.ncols, 'index': i, 'name': sheet.name}
            try:
                columns_dict[i] = sheet.row_values(0)
            except:
                pass
        return {'sheets': sheets_dict, 'nsheet': nsheets, 'names': names_dict, 'columns': columns_dict}

    def get_real_file(self, file_name):
        """
        生成存储excel文件的绝对路径，以及新文件名称
        加了文件已存在的check
        结果采用绝对路径+新文件名称，新文件名称
        generate to store real path of excel file
        if file exist dir, the file name rename file name + now(%Y-%m-%d-%H-%M-%S)
        the result is path file
        :param file_name: all path excel file
        :return: tuple
        result is contain real file, file name
        """
        now_date = get_now(format="%Y%m%d")
        real_store_dir = os.path.join(STORE_CACHE, now_date)
        if not os.path.exists(real_store_dir):
            mk_dirs(real_store_dir)
        _real_file = os.path.join(real_store_dir, file_name)
        # 文件已存在，加上时间戳
        if os.path.exists(_real_file):
            file_names = os.path.splitext(file_name)
            suffix = (file_names[1]).lower() if len(file_names) > 1 else ''
            new_file_name = '%s-%s%s' % (file_names[0], get_now(format="%Y-%m-%d-%H-%M-%S"), suffix)
            _real_file = os.path.join(real_store_dir, new_file_name)
            file_name = new_file_name
        return _real_file, file_name

    def get_split_store_dir(self, dir_name):
        """
        生成存储excel文件目录的绝对路径、相对路径
        generate split store path
        return real dir, relative dir
        :param dir_name: the store dir name
        :return: tuple
        result is contain real path, relative path
        """
        now_date_str = get_now(format="%Y%m%d")
        split_dir = '%s/%s' % (now_date_str, dir_name)
        real_store_dir = os.path.join(STORE_CACHE, split_dir)
        if os.path.exists(real_store_dir):
            split_dir = '%s/%s_%s' % (now_date_str, dir_name, get_now(format="%Y-%m-%d-%H-%M-%S"))
            real_store_dir = os.path.join(STORE_CACHE, split_dir)
            if not os.path.exists(real_store_dir):
                mk_dirs(real_store_dir)
        else:
            mk_dirs(real_store_dir)

        return real_store_dir

    def merge_openpyxl(self, new_name, file_list: list, **kwargs):
        """
        openpyxl: 不支持.xls（老版本excel）

        use openpyxl to merge excel, only it is to deal .xlsx formatter excel file
        the deal excel style is row values
        file_list format: {'file': f, 'sheets': set_sheet, 'nsheet': n}
            f: the all path excel file
            set_sheet: list, the select sheet list
            number sheet: sheet sum number
        :param file_list: excel list
        :param new_name: new excel name
        :param kwargs: multiple parameters, is dict type
            - blank: excel file to merge add some blank
        :return: json object
            status_id: result id, except status_id is 1 is success, others is failure
            message: the result messgae
            data:
                path: the new excel path + name, real path
                name: the new excel file name, not contain path
        """
        if not new_name:
            new_name = 'MERGE-%s%s' % (get_now(format="%Y-%m-%d-%H-%M-%S"), self.DEFAULT_PREFIX)
        if not file_list:
            return self.format_res(
                212, '合并文件列表不存在', {}
            )
        # check merge excel file list
        for _f in file_list:
            if not _f or not _f.get('file'): continue
            if not os.path.exists(_f.get('file')):
                return self.format_res(
                    302, '%s文件不存在，请重新上传' % _f.get('file'), {}
                )
            if os.path.splitext(_f.get('file'))[-1] == '.xls':
                return self.format_res(
                    312, '%s不支持.xls格式' % _f.get('file'), {}
                )
        if os.path.splitext(new_name)[-1] not in self.prefix_list:
            new_name = '%s%s' % (new_name, self.DEFAULT_PREFIX)
        # 分割行数,默认为0
        blank = kwargs.get('blank') or self.blank
        try:
            new_excel = openpyxl.Workbook()
            new_sheet = new_excel.active
            # new_sheet = new_excel.create_sheet(title='Sheet', index=0)
            # sum_row = 0
            for f in file_list:
                if not f or not f.get('file'): continue
                f_path = f.get('file')
                if not os.path.exists(f_path): continue

                sheet_index_list = f.get('sheets') if f.get('sheets') else [0]
                reader_excel = openpyxl.load_workbook(f_path)
                sheetnames = reader_excel.sheetnames
                max_nsheet = int(f.get('nsheet')) if f.get('nsheet') else len(sheetnames)
                for index in sheet_index_list:
                    index_int = int(index)
                    if index_int > max_nsheet: break
                    sheet = reader_excel.get_sheet_by_name(reader_excel.sheetnames[index_int])
                    # 行写入
                    for row in sheet.values:
                        new_sheet.append(row)
                    if blank > 0:
                        for _ in range(0, blank, 1): new_sheet.append([])
                    # 单元格写入
                    # sheet_row = sheet.max_row
                    # sheet_col = sheet.max_column
                    # for row in range(1, sheet_row+1, 1):
                    #     for col in range(1, sheet_col+1, 1):
                    #         # d = sheet.cell(row=row, column=col).value
                    #         new_sheet.cell(row=(sum_row+row),
                    #                        column=col,
                    #                        value=sheet.cell(row=row, column=col).value
                    #                        )
                    # sum_row += sheet_row + 1 + blank
            real_store_file, new_file_name = self.get_real_file(new_name)
            new_excel.save(real_store_file)
            return self.format_res(
                100, 'success', {'name': new_file_name, 'path': real_store_file}
            )
        except Exception as e:
            return self.format_res(
                998, str(e), {}
            )

    def merge_xlrw(self, new_name, file_list: list, **kwargs):
        """
        xlwt、xlrd: 表格行数限制65535

        use xlwt、xlrd to merge excel, it are to deal .xls、.xlsx formatter excel file
        the deal excel style is cell values(row + col)
        file_list format: 'file': f, 'sheets': set_sheet, 'nsheet': n}
            f: the all path excel file
            set_sheet: list, the select sheet list
            number sheet: sheet sum number
        :param file_list: excel list
        :param new_name: new excel file name
        :param kwargs: multiple parameters, is dict type
            - blank: excel file to merge add some blank
        :return: json object
            status_id: result id, except status_id is 1 is success, others is failure
            message: the result messgae
            data:
                path: the new excel path + name, real path
                name: the new excel file name, not contain path
        """
        if not new_name:
            new_name = 'MERGE-%s%s' % (get_now(format="%Y-%m-%d-%H-%M-%S"), self.DEFAULT_PREFIX)
        if not file_list:
            return self.format_res(
                212, '合并文件列表不存在', {}
            )
        # check merge excel file list
        for _f in file_list:
            if not _f or not _f.get('file'): continue
            if not os.path.exists(_f.get('file')):
                return self.format_res(
                    302, '%s文件不存在，请重新上传' % _f.get('file'), {}
                )
        if os.path.splitext(new_name)[-1] not in self.prefix_list:
            new_name = '%s%s' % (new_name, self.DEFAULT_PREFIX)
        # 分割行数,默认为0
        blank = kwargs.get('blank') if kwargs.get('blank') else self.blank
        try:
            new_excel = xlwt.Workbook(encoding='utf-8')
            new_sheet = new_excel.add_sheet('Sheet', cell_overwrite_ok=True)
            sum_row = 0
            for f in file_list:
                if not f or not f.get('file'): continue
                f_path = f.get('file')
                if not os.path.exists(f_path): continue

                sheet_index_list = f.get('sheets') if f.get('sheets') else [0]
                reader_excel = xlrd.open_workbook(f_path)
                max_nsheet = int(f.get('nsheet')) if f.get('nsheet') else len(reader_excel.sheet_names())
                for index in sheet_index_list:
                    index_int = int(index)
                    if index_int > max_nsheet: break
                    sheet = reader_excel.sheet_by_index(index_int)
                    sheet_row = sheet.nrows
                    sheet_col = sheet.ncols
                    for row in range(0, sheet_row, 1):
                        for col in range(0, sheet_col, 1):
                            new_sheet.write((sum_row+row), col, label=sheet.cell_value(row, col))
                    sum_row += (sheet_row + blank)
            real_store_file, new_file_name = self.get_real_file(new_name)
            new_excel.save(real_store_file)
            return self.format_res(
                100, 'success', {'name': new_file_name, 'path': real_store_file}
            )
        except Exception as e:
            return self.format_res(
                998, str(e), {}
            )

    def compress_zip(self, files, zip_name):
        """
        buildin function
        to use compress file
        :param files: the all real path file list
        :param zip_name: the new zip format file name
        :return: False or True
            False is failure
            True is success
        """
        if not files or not zip_name: return False
        zp = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
        for file in files:
            if not file or not os.path.split(file)[-1]: continue
            exe_f = os.path.split(file)[-1]
            zp.write(file, exe_f)
        zp.close()
        return zip_name

    def split_xlrw(self, file: str, **kwargs):
        """
        Use xlrd xlwt 处理.xls、.xlsx文件
        :param file: file path，包含项目路径，为存储的绝对路径
        :param kwargs:
            name：新文件名称 if not exist will to set current time to rename new file
            sheet: default 0 if not exist了，默认处理文件的第一个sheet
            type: 行拆分或者列拆分，default is 1 行拆分
            store: 存储方式，一表多Sheet或者多表一Sheet，default is 1 多表一Sheet
            rule: list formatter default 999 if not exist, 文件命名方式采用自增数字序列
            title: 新文件是否有标题，default is 1 有标题
        :return: json return

        type: 1行 2列
        store: 1多表一Sheet 2一表多Sheet
        999: 自增
        title: 1有标题 0无标题
        """
        if not file or not os.path.exists(file):
            return self.format_res(
                302, '文件不存在', {}
            )
        name = kwargs.get('name')
        if not name:
            name = 'SPLIT-%s' % get_now(format="%Y-%m-%d-%H-%M-%S")
        sheet = int(kwargs.get('sheet')) \
            if kwargs.get('sheet') else 0    # default is 0
        rc = str(kwargs.get('type')) \
            if kwargs.get('type') else '1'  # default is 1 行拆分 2列拆分
        store = str(kwargs.get('store')) \
            if kwargs.get('store') else '1'  # default is 1 多表一Sheet 2 一表多Sheet
        rule = kwargs.get('rule') \
            if kwargs.get('rule') else [self.DEFAULT_AUTO_ID]    # list formatter default is 9999 自增数字序列
        rule = [str(x) for x in rule]
        title = str(kwargs.get('title')) \
            if kwargs.get('title') else '1'  # default is 1 有标题
        if str(rc) not in EXCEL_NUM:
            return self.format_res(
                213, '请求参数split不合法', {}
            )
        if str(store) not in EXCEL_SPLIT_STORE:
            return self.format_res(
                213, '请求参数store不合法', {}
            )
        if str(title) not in BOOL:
            return self.format_res(
                213, '请求参数header不合法', {}
            )

        compress_name = name
        if os.path.splitext(compress_name)[-1] not in self.prefix_zip_list:
            compress_name = '%s%s' % (compress_name, self.DEFAULT_ZIP_PREFIX)
        reader_excel = xlrd.open_workbook(file)
        max_nsheet = len(reader_excel.sheet_names())
        index_int = int(sheet)
        if index_int >= max_nsheet:
            return self.format_res(
                229, '超出操作的sheet索引', {}
            )
        # 压缩 文件目录参数
        is_compress = False
        real_dir = ''

        # ======== start operation
        sheet = reader_excel.sheet_by_index(index_int)
        sheet_row = sheet.nrows
        sheet_col = sheet.ncols
        # type : 1 row 行拆分
        if rc == '1':
            if store == '1':    # store: 1 多表一Sheet
                real_dir = self.get_split_store_dir(name)
                for row in range(1, sheet_row, 1):
                    select_col = list()
                    # judge is or not 999 自增数字序列
                    if self.DEFAULT_AUTO_ID in rule:
                        select_col.append(str(row))
                    write_excel = xlwt.Workbook(encoding='utf-8')
                    new_sheet = write_excel.add_sheet('Sheet', cell_overwrite_ok=True)
                    for col in range(0, sheet_col, 1):
                        if title == '1':    # is or not write title
                            new_sheet.write(0, col, label=sheet.cell_value(0, col))
                        if str(col) in rule:
                            select_col.append(str(sheet.cell_value(row, col)))
                        new_sheet.write(1 if title == '1' else 0, col, label=sheet.cell_value(row, col))
                    new_excel_name = '%s%s' % ('_'.join(select_col), self.DEFAULT_PREFIX)
                    write_excel.save(os.path.join(real_dir, new_excel_name))
                is_compress = True
            elif store == '2':     # store: 2 一表多Sheet
                write_excel = xlwt.Workbook(encoding='utf-8')
                for row in range(1, sheet_row, 1):
                    select_col = list()
                    # judge is or not 999 自增数字序列
                    if self.DEFAULT_AUTO_ID in rule:
                        select_col.append(str(row))
                    for k, v in enumerate(sheet.row_values(row), 0):
                        if str(k) in rule: select_col.append(str(v))
                    new_sheet_name = '_'.join(select_col)
                    new_sheet = write_excel.add_sheet(new_sheet_name, cell_overwrite_ok=True)
                    for col in range(0, sheet_col, 1):
                        if title == '1':  # is or not write title
                            new_sheet.write(0, col, label=sheet.cell_value(0, col))
                        new_sheet.write(1 if title == '1' else 0, col, label=sheet.cell_value(row, col))
                    # end sheet
                new_excel_name = name
                if os.path.splitext(new_excel_name)[-1] not in self.prefix_list:
                    new_excel_name = '%s%s' % (new_excel_name, self.DEFAULT_PREFIX)
                real_store_file, new_excel_name = self.get_real_file(new_excel_name)
                write_excel.save(real_store_file)
                return self.format_res(
                    100, 'success', {'name': new_excel_name, 'path': real_store_file, 'compress': False}
                )
            else:
                return self.format_res(
                    213, '请求参数store不合法', {}
                )
        # type 2 col 列拆分
        elif rc == '2':
            if store == '1':    # store: 1 多表一Sheet
                real_dir = self.get_split_store_dir(name)
                for col in range(0, sheet_col, 1):
                    select_col = list()
                    # judge is or not 999 自增数字序列
                    if self.DEFAULT_AUTO_ID in rule:
                        select_col.append(str(col))
                    write_excel = xlwt.Workbook(encoding='utf-8')
                    new_sheet = write_excel.add_sheet('Sheet', cell_overwrite_ok=True)
                    for k, v in enumerate(sheet.row_values(0), 0):
                        if str(k) in rule: select_col.append(str(v))
                    # select_col.append(str(sheet.cell_value(0, col)))
                    if title == '1':   # is or not write title
                        new_sheet.write(0, 0, label=sheet.cell_value(0, col))
                    start_row = 1 if title == '1' else 0
                    for row in range(1, sheet_row, 1):
                        new_sheet.write(start_row, 0, label=sheet.cell_value(row, col))
                        start_row += 1
                    new_excel_name = '%s%s' % ('_'.join(select_col), self.DEFAULT_PREFIX)
                    write_excel.save(os.path.join(real_dir, new_excel_name))
                is_compress = True
            elif store == '2':     # store: 2 一表多Sheet
                write_excel = xlwt.Workbook(encoding='utf-8')
                for col in range(0, sheet_col, 1):
                    select_col = list()
                    # judge is or not 999 自增数字序列
                    if self.DEFAULT_AUTO_ID in rule:
                        select_col.append(str(col))
                    for k, v in enumerate(sheet.row_values(0), 0):
                        if str(k) in rule: select_col.append(str(v))
                    # select_col.append(str(sheet.cell_value(0, col)))
                    new_sheet_name = '_'.join(select_col)
                    new_sheet = write_excel.add_sheet(new_sheet_name, cell_overwrite_ok=True)
                    if title == '1':    # is or not write title
                        new_sheet.write(0, 0, label=sheet.cell_value(0, col))
                    start_row = 1 if title == '1' else 0
                    for row in range(1, sheet_row, 1):
                        new_sheet.write(start_row, 0, label=sheet.cell_value(row, col))
                        start_row += 1
                    # end sheet
                new_excel_name = name
                if os.path.splitext(new_excel_name)[-1] not in self.prefix_list:
                    new_excel_name = '%s%s' % (new_excel_name, self.DEFAULT_PREFIX)
                real_store_file, new_excel_name = self.get_real_file(new_excel_name)
                write_excel.save(real_store_file)
                return self.format_res(
                    100, 'success', {'name': new_excel_name, 'path': real_store_file, 'compress': False}
                )
        else:
            return self.format_res(
                213, '请求参数store不合法', {}
            )

        # start compress
        if is_compress:
            if not os.path.exists(real_dir) or not os.path.isdir(real_dir):
                return self.format_res(
                    230, '文件存储目录不存在', {}
                )
            try:
                zip_files = [os.path.join(real_dir, x) for x in os.listdir(real_dir)]
                is_ok = self.compress_zip(files=zip_files,
                                          zip_name=os.path.join(real_dir, compress_name))
                if is_ok:
                    return self.format_res(
                        100,
                        'success',
                        {'name': compress_name, 'path': os.path.join(real_dir, compress_name), 'nfile': len(zip_files), 'compress': True}
                    )
            except:
                return self.format_res(
                    234, '压缩文件有误', {}
                )

        return self.format_res(
            999, '暂无其他处理方式', {}
        )

